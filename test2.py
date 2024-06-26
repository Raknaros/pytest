import pandas as pd
from openpyxl import load_workbook

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

# Cargar el libro de Excel
workbook_path = 'C:/Users/Giu/OneDrive/facturacion/cuadros.xlsx'
workbook = load_workbook(workbook_path)

# Crear un diccionario para almacenar los resultados
resultados = {}

# Iterar sobre todas las hojas del libro
for sheet_name in workbook.sheetnames:
    # Seleccionar la hoja actual
    sheet = workbook[sheet_name]

    # Recuperar la primera fila
    primera_fila = [cell.value for cell in sheet[1][:6]]

    # Crear una lista de listas con el resto del contenido de la hoja
    resto_filas = [[cell.value for cell in row[:25]] for row in sheet.iter_rows(min_row=2)]

    # Convertir el resto del contenido en un DataFrame
    df = pd.DataFrame(resto_filas[1:], columns=resto_filas[0])

    # Almacenar los resultados en el diccionario
    resultados[sheet_name] = {'primera_fila': primera_fila, 'dataframe': df}

# Ejemplo de acceso a los resultados
for sheet_name, data in resultados.items():
    print(f"Hoja: {sheet_name}")
    print("Primera fila:", data['primera_fila'])
    print("DataFrame:")
    print(data['dataframe'])
